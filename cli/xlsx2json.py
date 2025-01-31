#!/usr/bin/env python
import argparse
import datetime
import json
import logging
from json import JSONEncoder

import openpyxl
import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'
HEADERS_KEY_ROW = 'header_row'
CAST_KEY_NAME = 'cast'


class DateTimeEncoder(JSONEncoder):
    # Override the default method
    def default(self, obj):
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()


class XlsxParser:
    """
    Base parser for Excel file for the fields from worksheets defined in a configuration file.
    It implements the base functionality for opening and validating the spreadsheet.
    """

    def __init__(self, xlsx_filename, conf_filename):
        """
        Constructor
        :param xlsx_filename: Excel file path
        :type xlsx_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xlsx_conf = yaml.safe_load(conf_file)
        try:
            self.workbook = load_workbook(xlsx_filename, read_only=True)
        except Exception as e:
            logging.error('Error loading %s', xlsx_filename)
            raise e
        self.worksheets = None
        self._active_worksheet = None
        self.row_offset = {}
        self.headers = {}
        self.valid = None

    @property
    def active_worksheet(self):
        return self._active_worksheet

    @active_worksheet.setter
    def active_worksheet(self, worksheet):
        if self.worksheets is None:
            self.valid_worksheets()
        if worksheet not in self.worksheets:
            raise ValueError('Worksheet ' + worksheet + ' is not a valid worksheet!')

        self._active_worksheet = worksheet

    def valid_worksheets(self):
        """
        Get the list of the names of worksheets which have all the configured required headers
        :return: list of valid worksheet names in the Excel file
        :rtype: list
        """
        if self.worksheets is not None:
            return self.worksheets

        self.worksheets = []
        sheet_titles = self.workbook.sheetnames

        for title in self.xlsx_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            header_row = self.xlsx_conf[title].get(HEADERS_KEY_ROW, 1)
            if worksheet.max_row < header_row + 1:
                continue
            # Check required headers are present
            self.headers[title] = [cell.value if cell.value is None else cell.value.strip()
                                   for cell in worksheet[header_row]]
            required_headers = self.xlsx_conf[title].get(REQUIRED_HEADERS_KEY_NAME, [])
            if set(required_headers) <= set(self.headers[title]):  # issubset
                self.worksheets.append(title)
            else:
                missing_headers = set(required_headers) - set(self.headers[title])
                self.valid = False
                raise ValueError('Worksheet '+title+' does not have all the required headers!: '+ ','.join(missing_headers) )

        return self.worksheets

    def is_valid(self):
        """
        Check that is all the worksheets contain required headers
        :return: True if all the worksheets contain required headers. False otherwise
        :rtype: bool
        """
        if self.valid is None:
            self.valid = True
            self.valid_worksheets()

        return self.valid

    @staticmethod
    def cast_value(value, type_name):
        # Do not cast None values
        if type_name and value is not None:
            if type_name == 'string':
                return str(value)
        return value

    @staticmethod
    def trim_value(value):
        """Remove whitespace from the start and end of cells"""
        if isinstance(value, str):
            return value.strip()
        return value

    def base_row_offset(self, worksheet):
        return self.xlsx_conf[worksheet].get(HEADERS_KEY_ROW, 1)

    def get_rows(self):
        """
        Retrieve all the data rows.
        :return: list of hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: list
        """
        worksheet = self.active_worksheet
        if worksheet is None:
            logging.warning('No worksheet is specified!')
            return None

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = self.base_row_offset(worksheet)
        self.row_offset[worksheet] += 1

        required_headers = self.xlsx_conf[worksheet].get(REQUIRED_HEADERS_KEY_NAME, {})
        optional_headers = self.xlsx_conf[worksheet].get(OPTIONAL_HEADERS_KEY_NAME, {})

        rows = []

        for row in self.workbook[worksheet].iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            has_notnull = False
            for header in list(required_headers.keys()) + list(optional_headers.keys()):
                header_index = num_cells
                if header in self.headers[worksheet]:
                    header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    data[header] = None
                    continue

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True
                data[header] = self.trim_value(self.cast_value(
                    cell.value, self.xlsx_conf[worksheet].get(CAST_KEY_NAME, {}).get(header)
                ))

            if has_notnull:
                data['row_num'] = self.row_offset[worksheet]
                rows.append(data)

            # no data on this row, continue to next
            self.row_offset[worksheet] += 1

        return rows

    def json(self, output_json_file):
        json_data = {}
        for title in self.xlsx_conf[WORKSHEETS_KEY_NAME]:
            json_data[self.xlsx_conf[WORKSHEETS_KEY_NAME][title]] = []
            self.active_worksheet = title
            for row in self.get_rows():
                # Remove the row number
                row.pop('row_num')
                # Remove any None and translate header name
                json_data[self.xlsx_conf[WORKSHEETS_KEY_NAME][title]].append(
                    {self.translate_header(title, k): v for k, v in row.items() if v is not None}
                )

        with open(output_json_file, 'w') as open_file:
            json.dump(json_data, open_file, cls=DateTimeEncoder)

    def translate_header(self, title, header):
        if header in self.xlsx_conf[title].get(REQUIRED_HEADERS_KEY_NAME, {}):
            return self.xlsx_conf[title][REQUIRED_HEADERS_KEY_NAME][header]
        if header in self.xlsx_conf[title].get(OPTIONAL_HEADERS_KEY_NAME, {}):
            return self.xlsx_conf[title][OPTIONAL_HEADERS_KEY_NAME][header]
        logging.warning(f'Header {header} in {title} sheet does not have translation in the config file. Leave it as is')
        return header


def create_xls_template_from_yaml(xlsx_filename, conf_filename):
    """
    Create a XLS empty template with the expected worksheet and header based on the configuration provided
    """
    with open(conf_filename, 'r') as conf_file:
        xlsx_conf = yaml.safe_load(conf_file)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title in xlsx_conf[WORKSHEETS_KEY_NAME]:
        # Create worksheet
        workbook.create_sheet(title)
        # For each Worksheet create the header row
        header_row = xlsx_conf[title].get(HEADERS_KEY_ROW, 1)
        header_col_index = 0
        for header_name in xlsx_conf[title].get(REQUIRED_HEADERS_KEY_NAME, []):
            header_col_index += 1
            cell = workbook[title].cell(column=header_col_index, row=header_row, value=header_name)
            cell.font = openpyxl.styles.Font(bold=True)
        for header_name in xlsx_conf[title].get(OPTIONAL_HEADERS_KEY_NAME, []):
            header_col_index += 1
            workbook[title].cell(column=header_col_index, row=header_row, value=header_name)
    workbook.save(xlsx_filename)


def main():
    arg_parser = argparse.ArgumentParser(
        description='Convert an xlsx spreadsheet containing expected sheets to json prior to validation')
    arg_parser.add_argument('--metadata_json', required=True, dest='metadata_json', help='Path to output EVA metadata json file')
    arg_parser.add_argument('--metadata_xlsx', required=True, dest='metadata_xlsx', help='EVA metadata Excel file')
    arg_parser.add_argument('--conversion_configuration', dest='conversion_configuration',
                            help='Configuration file describing the expected content of the Excel spreadsheet')

    args = arg_parser.parse_args()
    parser = XlsxParser(args.metadata_xlsx, args.conversion_configuration)
    parser.json(args.metadata_json)


if __name__ == "__main__":
    main()
